import openpyxl
from openpyxl.styles import PatternFill
import re
import sys 

# --- Configuration ---
FILE_PATH = "Cobden_JUNE_2025_Lab_and_Insitu_Data_V1 (1).xlsx"
OUTPUT_PATH = "Cobden_Processed.xlsx"
SHEET_NAME_KEYWORD = "Attachment 3"
THRESHOLD_START_ROW = 16
THRESHOLD_END_ROW = 21
DATA_START_ROW = 23 # Assumption based on Row 22 being headers in inspection

# --- Logic ---
def parse_threshold(text):
    """
    Parses a cell string to find threshold logic.
    Returns: (value, operator_type) or None
    operator_type: 'GT' (> value), 'LT' (< value), 'GT_MAX' (Range, > Max)
    """
    text = str(text).strip()
    if not text:
        return None
    
    # Check for inequalities
    if '<' in text:
        # Extract number
        matches = re.findall(r"[-+]?\d*\.\d+|\d+", text)
        if matches:
            return (float(matches[-1]), 'LT') # Use last number found (e.g. < 5.5)
            
    if '>' in text:
        matches = re.findall(r"[-+]?\d*\.\d+|\d+", text)
        if matches:
            return (float(matches[-1]), 'GT')
            
    # Check for Range (e.g. 6.5-8.5)
    # Detect hyphen between numbers
    range_match = re.search(r"(\d*\.?\d+)\s*-\s*(\d*\.?\d+)", text)
    if range_match:
        # usually ranges implies acceptance inside. So "Above" -> > Max.
        # User explicitly said "Above", so we color if > Max.
        # Note: If it's pH 6.0-8.5, "Above" fails if > 8.5. 
        # (It ignores < 6.0 failure, but follows user strict "Above" instruction)
        return (float(range_match.group(2)), 'GT') 

    # Simple number
    # Remove #footnotes (e.g. 6.0-8.5#6 -> handled by range, but simple 500#2 -> 500)
    # Just extracting first float
    # If text is simple number "8.03", treat as Limit (GT)
    try:
        # If it looks like a number
        val = float(text)
        return (val, 'GT')
    except:
        # Regex for number
        matches = re.findall(r"[-+]?\d*\.\d+|\d+", text)
        if matches:
             # Take the first number found (ignores footnotes like #2)
             return (float(matches[0]), 'GT')
        
    return None

def main():
    print(f"Loading {FILE_PATH}...")
    try:
        # Load workbook (need data_only=True to read values? 
        # But we need styles from standard load. 
        # Solution: Load twice. Once for values/styles (standard), 
        # openpyxl with data_only=False doesn't evaluate formulas.
        # Best: Inspect values using data_only=True logic if needed.
        # For this file, raw numbers likely exist.
        wb = openpyxl.load_workbook(FILE_PATH)
        
        # Find Sheet
        target_sheet = next((s for s in wb.sheetnames if SHEET_NAME_KEYWORD in s), None)
        if not target_sheet:
            print(f"Sheet with '{SHEET_NAME_KEYWORD}' not found.")
            return
        
        ws = wb[target_sheet]
        print(f"Processing Sheet: {ws.title}")

        # 0. Find pH Column Dynamically (to skip it as requested)
        # Scan header row (Row 11 based on previous inspection) for "pH"
        ph_col_index = -1
        header_row = 11
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val and "pH" in str(val):
                ph_col_index = col_idx
                print(f"Detected pH column at Index {ph_col_index} (will skip coloring for this col)")
                break

        # 1. Parse Thresholds (Rows 16-21)
        # Structure: thresholds[col_index] = list of {val, op, fill}
        col_thresholds = {}
        
        # Determine max column
        max_col = ws.max_column
        
        for row_idx in range(THRESHOLD_START_ROW, THRESHOLD_END_ROW + 1):
            row_cells = list(ws.rows)[row_idx-1] # 0-indexed list access
            
            for col_idx in range(5, max_col + 1): # Start from Col 6 (Index 5? No 1-based in iter)
                # openpyxl cell access is 1-based
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # We need the value. If formula, this might be None if not cached.
                # Assuming static values for thresholds matching inspection.
                val = cell.value
                if not val:
                    continue
                
                parsed = parse_threshold(val)
                if not parsed:
                    continue
                    
                limit_val, op = parsed
                
                # Get Style
                # We need the PatternFill object or copy it
                # Note: Theme colors are tricky. We will try to copy the raw fill.
                source_fill = cell.fill
                
                if col_idx not in col_thresholds:
                    col_thresholds[col_idx] = []
                    
                col_thresholds[col_idx].append({
                    'limit': limit_val,
                    'op': op,
                    'fill': source_fill, # We will re-apply this object
                    'row_desc': row_idx
                })

        print(f"Parsed thresholds for {len(col_thresholds)} columns.")

        # 2. Iterate Data Rows
        # Find actual data start? Scanning for a Date in Col 2?
        # User manual verification said 'Meaning 6 rows... below these start the data'
        # Inspection showed Row 22 empty/header. Let's start Row 23.
        
        data_rows_processed = 0
        cells_colored = 0
        
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            # Optimization: Stop if Row is empty?
            # Check date column (Col 2)
            if not ws.cell(row=row_idx, column=2).value:
                 # Be careful not to stop prematurely if just one gap
                 pass

            for col_idx in range(5, max_col + 1):
                # SKIP pH Column (Dynamic check)
                if col_idx == ph_col_index:
                    continue

                if col_idx not in col_thresholds:
                    continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                
                if val is None:
                    continue
                    
                # Ensure val is numeric
                try:
                    num_val = float(val)
                except:
                    continue
                    
                # Check Thresholds
                # Strategy: Identify ALL matching thresholds.
                # Use Priority: Highest Limit Value wins (assuming higher = more severe).
                
                matches = []
                
                for thresh in col_thresholds[col_idx]:
                    limit = thresh['limit']
                    op = thresh['op']
                    
                    is_match = False
                    if op == 'GT': # Limit is Max. Trigger if > Limit
                        if num_val > limit:
                            is_match = True
                    elif op == 'LT': 
                        if num_val > limit:
                            is_match = True
                            
                    if is_match:
                        matches.append(thresh)
                        
                if matches:
                    # Pick the best match. 
                    # If we have [1000, 2000]. Data=3000.
                    # We want 2000 (usually red/more severe).
                    best_match = max(matches, key=lambda x: x['limit'])
                    
                    active_fill = best_match['fill']
                    
                    # Apply
                    # Copying fill is safer than assigning ref if cross-workbook, 
                    # but same workbook is okay. 
                    # Ideally create new PatternFill with same props to avoid linking.
                    new_fill = PatternFill(
                        patternType=active_fill.patternType,
                        fgColor=active_fill.fgColor,
                        bgColor=active_fill.bgColor
                    )
                    cell.fill = new_fill
                    cells_colored += 1

            
            data_rows_processed += 1
            
        print(f"Processed {data_rows_processed} rows. Colored {cells_colored} cells.")
        wb.save(OUTPUT_PATH)
        print(f"Saved to {OUTPUT_PATH}")

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
