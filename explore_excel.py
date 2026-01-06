import openpyxl
import re

FILE_PATH = "Cobden_JUNE_2025_Lab_and_Insitu_Data_V1 (1).xlsx"
SHEET_NAME_KEYWORD = "Attachment 3"
TDS_COL_IDX = 15

def parse_threshold(text):
    text = str(text).strip()
    if not text: return None, None
    if '<' in text:
        matches = re.findall(r"[-+]?\d*\.\d+|\d+", text)
        if matches: return float(matches[-1]), 'LT'
    if '>' in text:
        matches = re.findall(r"[-+]?\d*\.\d+|\d+", text)
        if matches: return float(matches[-1]), 'GT'
    range_match = re.search(r"(\d*\.?\d+)\s*-\s*(\d*\.?\d+)", text)
    if range_match:
        return float(range_match.group(2)), 'GT' 
    try:
        val = float(text)
        return val, 'GT'
    except:
        return None, None

try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    target_sheet = next((s for s in wb.sheetnames if SHEET_NAME_KEYWORD in s), None)
    ws = wb[target_sheet]
    
    print(f"--- Inspecting TDS (Col {TDS_COL_IDX}) thresholds ---")
    thresholds = []
    for r in range(16, 22):
        cell = ws.cell(row=r, column=TDS_COL_IDX)
        val = cell.value
        parsed_val, op = parse_threshold(val)
        
        # Get Color Info
        color_info = "None"
        if cell.fill and cell.fill.start_color:
            sc = cell.fill.start_color
            if sc.type == 'rgb': color_info = f"RGB:{sc.rgb}"
            else: color_info = f"Type:{sc.type}"

        print(f"Row {r}: '{val}' -> Parsed: {op} {parsed_val} | Color: {color_info}")
        if parsed_val:
            thresholds.append({'r':r, 'lim':parsed_val, 'op':op})

    print("\n--- Logic Simulation ---")
    sim_vals = [900, 1500, 2500]
    for v in sim_vals:
        print(f"Data Value: {v}")
        matches = []
        for t in thresholds:
            if t['op'] == 'GT' and v > t['lim']:
                matches.append(t)
        
        if not matches:
             print("  -> No Match")
        else:
             print(f"  -> Matches: {[m['lim'] for m in matches]}")
             # Current Code Logic: Last one wins
             print(f"  -> Winner (Current Last-One-Wins): {matches[-1]['lim']} (Row {matches[-1]['r']})")
             # Proposed Logic: Highest Limit wins
             best = max(matches, key=lambda x: x['lim'])
             print(f"  -> Winner (Proposed Highest-Val-Wins): {best['lim']} (Row {best['r']})")

except Exception as e:
    print(e)
