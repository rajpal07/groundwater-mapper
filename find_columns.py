import openpyxl

FILE_PATH = "Cobden_JUNE_2025_Lab_and_Insitu_Data_V1 (1).xlsx"
SHEET_NAME_KEYWORD = "Attachment 3"

try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    target_sheet = next((s for s in wb.sheetnames if SHEET_NAME_KEYWORD in s), None)
    ws = wb[target_sheet]
    
    # Inspect Row 11 (Header Row based on prev output)
    print("--- Searching Headers in Row 11 ---")
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=11, column=col).value
        headers.append((col, str(val).strip() if val else ""))
        
    print("Columns Found:")
    for col_idx, name in headers:
        if "Solid" in name or "Chloride" in name or "pH" in name:
            print(f"  Col {col_idx}: '{name}'")

except Exception as e:
    print(e)
