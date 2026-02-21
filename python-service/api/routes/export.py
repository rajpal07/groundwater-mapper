"""
Smart Export routes for Excel export with conditional formatting.
"""

import logging
from typing import Optional, List

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import Response
import pandas as pd
import io
import base64

from ..models import SmartExportResponse
from ..services.excel_parser import excel_parser
from ..auth import get_current_user_optional

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["Export"])


@router.post("/smart-excel", response_model=SmartExportResponse)
async def smart_excel_export(
    file: UploadFile = File(..., description="Excel file to process"),
    parameter: str = Form(..., description="Column name for the parameter"),
    thresholds: str = Form(..., description="JSON string of threshold values, e.g., '[10, 20, 30]'"),
    colors: Optional[str] = Form(None, description="JSON string of colors for each range"),
    sheet_name: Optional[str] = Form(None, description="Sheet name to process"),
    output_filename: Optional[str] = Form(None, description="Output filename"),
    user: Optional[dict] = Depends(get_current_user_optional)
) -> SmartExportResponse:
    """
    Export Excel file with conditional formatting based on thresholds.
    
    This endpoint:
    1. Parses the Excel file
    2. Applies conditional formatting based on threshold values
    3. Returns a new Excel file with color-coded cells
    
    Thresholds define the boundaries for different color zones.
    For example, thresholds [10, 20, 30] create 4 zones:
    - < 10: Green (low)
    - 10-20: Yellow (moderate)
    - 20-30: Orange (high)
    - > 30: Red (critical)
    """
    import json
    
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        raise HTTPException(
            status_code=400, 
            detail="Invalid file type. Only .xlsx, .xls, and .csv files are supported."
        )
    
    # Parse thresholds
    try:
        threshold_values = json.loads(thresholds)
        if not isinstance(threshold_values, list):
            raise ValueError("Thresholds must be a list")
        threshold_values = [float(t) for t in threshold_values]
        threshold_values.sort()
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid thresholds format: {str(e)}. Expected JSON array like '[10, 20, 30]'"
        )
    
    # Parse colors if provided
    color_values = None
    if colors:
        try:
            color_values = json.loads(colors)
            if not isinstance(color_values, list):
                raise ValueError("Colors must be a list")
        except (json.JSONDecodeError, ValueError) as e:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid colors format: {str(e)}. Expected JSON array like '[\"#00FF00\", \"#FFFF00\"]'"
            )
    
    # Read file content
    try:
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="Empty file uploaded")
    except Exception as e:
        logger.error(f"Error reading file: {e}")
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")
    
    # Parse Excel file
    try:
        df = excel_parser.parse_file(content, filename=file.filename, sheet_name=sheet_name)
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise HTTPException(status_code=422, detail=f"Error parsing Excel file: {str(e)}")
    
    # Validate parameter column
    if parameter not in df.columns:
        raise HTTPException(
            status_code=400, 
            detail=f"Parameter column '{parameter}' not found. Available columns: {list(df.columns)}"
        )
    
    # Generate Excel with conditional formatting
    try:
        excel_base64, stats = generate_conditional_excel(
            df=df,
            parameter=parameter,
            thresholds=threshold_values,
            colors=color_values
        )
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")
    
    return SmartExportResponse(
        success=True,
        message="Excel file generated with conditional formatting",
        filename=output_filename or f"processed_{file.filename}",
        parameter=parameter,
        thresholds=threshold_values,
        statistics=stats,
        excel_base64=excel_base64
    )


def generate_conditional_excel(
    df: pd.DataFrame,
    parameter: str,
    thresholds: List[float],
    colors: Optional[List[str]] = None
) -> tuple:
    """
    Generate Excel file with conditional formatting.
    
    Returns:
        tuple: (base64_encoded_excel, statistics_dict)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    import numpy as np
    
    # Default colors for threshold ranges (green to red gradient)
    if not colors:
        colors = [
            "#00FF00",  # Green (lowest)
            "#90EE90",  # Light green
            "#FFFF00",  # Yellow
            "#FFA500",  # Orange
            "#FF0000",  # Red (highest)
        ]
    
    # Ensure we have enough colors for all ranges
    num_ranges = len(thresholds) + 1
    while len(colors) < num_ranges:
        colors.append("#FF0000")  # Default to red for extra ranges
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Data"
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Find parameter column index
    param_col_idx = list(df.columns).index(parameter) + 1
    
    # Write data with conditional formatting
    value_stats = {
        "total_rows": len(df),
        "ranges": []
    }
    
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply conditional formatting to parameter column
            if col_idx == param_col_idx and pd.notna(value):
                try:
                    num_value = float(value)
                    color_idx = get_color_index(num_value, thresholds)
                    fill = PatternFill(
                        start_color=colors[color_idx].lstrip('#'),
                        end_color=colors[color_idx].lstrip('#'),
                        fill_type='solid'
                    )
                    cell.fill = fill
                except (ValueError, TypeError):
                    pass
    
    # Calculate statistics for each range
    param_values = pd.to_numeric(df[parameter], errors='coerce')
    for i in range(len(thresholds) + 1):
        if i == 0:
            lower = float('-inf')
            upper = thresholds[0]
        elif i == len(thresholds):
            lower = thresholds[-1]
            upper = float('inf')
        else:
            lower = thresholds[i - 1]
            upper = thresholds[i]
        
        mask = (param_values >= lower) & (param_values < upper)
        count = mask.sum()
        
        value_stats["ranges"].append({
            "range": f"{lower:.2f} - {upper:.2f}" if upper != float('inf') else f">= {lower:.2f}",
            "count": int(count),
            "percentage": round(count / len(df) * 100, 2) if len(df) > 0 else 0,
            "color": colors[i]
        })
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Parameter"
    ws_summary['B1'] = parameter
    ws_summary['A2'] = "Total Rows"
    ws_summary['B2'] = len(df)
    ws_summary['A3'] = "Thresholds"
    ws_summary['B3'] = str(thresholds)
    
    # Add range statistics
    ws_summary['A5'] = "Range Statistics"
    ws_summary['A6'] = "Range"
    ws_summary['B6'] = "Count"
    ws_summary['C6'] = "Percentage"
    ws_summary['D6'] = "Color"
    
    for i, range_stat in enumerate(value_stats["ranges"], 7):
        ws_summary[f'A{i}'] = range_stat["range"]
        ws_summary[f'B{i}'] = range_stat["count"]
        ws_summary[f'C{i}'] = f"{range_stat['percentage']}%"
        ws_summary[f'D{i}'] = range_stat["color"]
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Encode to base64
    excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
    
    return excel_base64, value_stats


def get_color_index(value: float, thresholds: List[float]) -> int:
    """
    Get the color index for a value based on thresholds.
    
    Args:
        value: The value to check
        thresholds: Sorted list of threshold values
    
    Returns:
        int: Index of the color to use
    """
    for i, threshold in enumerate(thresholds):
        if value < threshold:
            return i
    return len(thresholds)
