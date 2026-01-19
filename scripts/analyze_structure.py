import openpyxl
import pandas as pd

def inspect_file(filepath):
    print(f"--- Inspecting {filepath} ---")
    try:
        # Check Sheet Names
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print("Sheets:", wb.sheetnames)
        
        # Check for keywords in first 50 rows of first sheet (or relevant sheet)
        for sheet in wb.sheetnames:
            print(f"Scanning sheet: {sheet}")
            ws = wb[sheet]
            found_att3 = False
            found_well = False
            
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
                row_str = str(row)
                if "Attachment 3" in row_str:
                    print(f"  [Row {i+1}] Found 'Attachment 3'")
                    found_att3 = True
                if "Well ID" in row_str:
                    print(f"  [Row {i+1}] Found 'Well ID'")
                    found_well = True
                    
            if not found_att3:
                print("  X 'Attachment 3' NOT found in first 50 rows.")
            if not found_well:
                print("  X 'Well ID' NOT found in first 50 rows.")
                
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    inspect_file("Cobden_JUNE_2025_Lab_and_Insitu_Data_V1 (1).xlsx")
    print("\n")
    inspect_file("STANHOPE_JAN-25.xlsx")
    print("\n")
    inspect_file("DARNUM_NOV-25.xlsx")
